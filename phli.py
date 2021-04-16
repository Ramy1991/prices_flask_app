from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from lxml import html
import os
import time
from datetime import datetime

executable_path = r'C:\Users\ramyg\Downloads\chromedriver_win32 (1)\chromedriver.exe'
os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
driver = webdriver.Chrome(executable_path=executable_path)

wb = load_workbook(filename=r'C:\Users\ramyg\Desktop\pph.xlsx')
ws = wb.active
i = 2

for row in ws.rows:
    link = ws['a' + str(i)].value
    driver.get(link)
    time.sleep(1)
    page = driver.page_source
    tree = html.fromstring(page)
    des = tree.xpath('/html/body/div[3]/div/div[5]/div[2]/div[1]/div/div/div/div[2]/section/div/div[1]//h3//text()')
    des2 = tree.xpath('//*[@id="see-all-benefits"]/div/section/div/ul//li//text()')
    ws.cell(row=i, column=2).value = '\n'.join(des)
    ws.cell(row=i, column=3).value = '\n'.join(des2)
    wb.save(r'C:\Users\ramyg\Desktop\pph.xlsx')
    print(des)
    i += 1
