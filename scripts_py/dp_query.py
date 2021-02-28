# import mysql.connector
# from mysql.connector import errorcode
#
# # Obtain connection string information from the portal
# try:
#     conn = mysql.connector.connect(user="admin1991@prices-app", password="Ramy1991",
#                                    host="prices-app.mysql.database.azure.com", port=3306, database="prices_test",
#                                    ssl_ca="E:\Pycharm\prices_flask\scripts_py\DigiCertGlobalRootG2.crt.pem",
#                                    ssl_verify_cert=True)
#
# except mysql.connector.Error as err:
#     if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
#         print("Something is wrong with the user name or password")
#     elif err.errno == errorcode.ER_BAD_DB_ERROR:
#         print("Database does not exist")
#     else:
#         print(err)
# else:
#     cursor = conn.cursor()
#     cursor.execute("SELECT COUNT(*) FROM products")
#     query = cursor.fetchall()
#     print(query)
#     #
#     # for x in cursor:
#     #     print(x)

import mysql.connector
from mysql.connector import errorcode
from openpyxl import Workbook, load_workbook
import re
import uuid
from super_cate_mod import super_care_func


# Obtain connection string information from the portal

def remove_special_characters(string):
    replaced = re.sub(r"['|\\\|,|\s\/|-|.|@|$|%|(|)|_|&|+|=|^|{|}|<|>|\"|!|?|:|;]+", "-", string)
    return re.sub(r"[-]+", "-", replaced)


try:
    conn = mysql.connector.connect(user="root", password="",
                                   host="localhost", port=3306, database="test",
                                   )

except mysql.connector.Error as err:
    if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Something is wrong with the user name or password")
    elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Database does not exist")
    else:
        print(err)
else:
    wb = load_workbook(filename='KSA_TO_UPLOAD.xlsx')
    ws = wb.active
    i = 2
    f = ws.max_row + 1
    max_r = ws.max_row

    for row in ws.rows:
        Ean = ws['a' + str(i)].value
        Country = ws['b' + str(i)].value
        Item_Type_EN = ws['c' + str(i)].value
        Item_Type_AR = ws['e' + str(i)].value
        title_en = ws['f' + str(i)].value
        title_ar = ws['g' + str(i)].value
        Brand_en = ws['h' + str(i)].value
        Brand_ar = ws['i' + str(i)].value
        Description_en = ws['j' + str(i)].value
        Description_ar = ws['k' + str(i)].value
        Item_Specs_en = ws['l' + str(i)].value
        Item_Specs_ar = ws['m' + str(i)].value
        images = ws['n' + str(i)].value
        Product_Direct_Link_en = ws['o' + str(i)].value
        Product_Direct_Link_ar = ws['p' + str(i)].value
        Price_EG = ws['r' + str(i)].value
        UPCs = ws['x' + str(i)].value
        Sold_Out = ws['y' + str(i)].value

        title_en_ = re.sub("'", "''", title_en)
        title_ar_ = re.sub("'", "''", title_ar)

        Brand_en_ = re.sub("'", "''", Brand_en)
        Brand_ar_ = re.sub("'", "''", Brand_ar)

        Description_en_ = re.sub("'", "''", Description_en)
        Description_ar_ = re.sub("'", "''", Description_ar)
        Item_Specs_en_ = re.sub("'", "''", Item_Specs_en)
        Item_Specs_ar_ = re.sub("'", "''", Item_Specs_ar)
        Item_Type_EN_ = re.sub("'", "''", Item_Type_EN)
        Item_Type_AR_ = re.sub("'", "''", Item_Type_AR)

        uic = "A0{}".format(str(uuid.uuid4()).upper()[:8])

        URL_EN = remove_special_characters(title_en)
        URL_AR = remove_special_characters(title_ar)

        Cate_URL_EN = remove_special_characters(Item_Type_EN_)
        Cate_URL_AR = remove_special_characters(Item_Type_AR_)

        Super_Category = str(super_care_func(re.sub(r"\s+", " ", Item_Type_EN).strip()))
        if ''.join(Super_Category) == 'None':
            sub_category_en = "not_found"
            sub_category_ar = "not_found"
            sub_category_URL_EN = ''
        else:
            Super_Category_ = Super_Category.split('@')
            sub_category_en = Super_Category_[0]
            sub_category_ar = Super_Category_[1]
            sub_category_URL_EN = re.sub("'", "''", remove_special_characters(sub_category_en))

        link_en = '/'.join([Cate_URL_EN, URL_EN, uic])
        link_ar = '/'.join([Cate_URL_AR, URL_AR, uic])
        # print(Product_Direct_Link_en)
        query = (
            "INSERT INTO Products (Unique_Product_Code, Country, sub_category_URL_EN, sub_category_en, sub_category_ar, Item_Type_EN, Item_Type_AR, Title_EN, Title_AR, Brand_EN, Brand_AR, Description_EN, Description_AR, Item_Specs_en, Item_Specs_ar, Images_URL, Product_Direct_Link_EN, Product_Direct_Link_AR, Price_EG, Item_UPC, Sold_Out, Affiliate_Link, URL_EN, URL_AR, UIC, Cate_URL_EN, Cate_URL_AR, link_en, link_ar, added_to_specs, Website_Name)"
            "VALUES (%s, %s, %s,%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'No', 'souq.com')")

        data = (
            Ean, Country, sub_category_URL_EN, sub_category_en, sub_category_ar, Item_Type_EN_, Item_Type_AR_,
            title_en_,
            title_ar_, Brand_en_, Brand_ar_, Description_en_, Description_ar_, Item_Specs_en_, Item_Specs_ar_, images,
            Product_Direct_Link_en, Product_Direct_Link_ar, Price_EG, UPCs, Sold_Out, 'aff', URL_EN, URL_AR, uic,
            Cate_URL_EN, Cate_URL_AR, link_en, link_ar)

        cursor = conn.cursor()
        try:
            cursor.execute(query, data)
            conn.commit()

        except mysql.connector.errors.IntegrityError as err:
            print(err)

        print(cursor.rowcount, "record inserted.")
        i = i + 1
