from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from lxml import html
import os
import time
from datetime import datetime

executable_path = r"C:\Users\ramyg\Downloads\chromedriver_win32 (1)\chromedriver.exe"
os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
driver = webdriver.Chrome(executable_path=executable_path)
driver.get('https://mzadqatar.com/mzadqater_admin/login')

driver.find_element_by_xpath('/html/body/div[3]/div/div/div/form/div/div[2]/input').send_keys(
    'ramy.zaghloul@mzadqatar.com')
driver.find_element_by_xpath('/html/body/div[3]/div/div/div/form/div/div[3]/input').send_keys('MkE83Q@u')
driver.find_element_by_xpath('/html/body/div[3]/div/div/div/form/div/div[4]/button').click()

driver.get('https://mzadqatar.com/mzadqater_admin/lots/create-lot')

driver.find_element_by_xpath("//*[@id='choose_category']/option[contains(text(),'Car plates')]").click()
wb = load_workbook(filename=r"C:\Users\ramyg\Downloads\Bidding Data Sheet (2).xlsx")
ws = wb.active

last_col = ws.max_column
col_num = 1
mandatory_att = []
for col in ws.iter_cols(min_row=1, max_col=last_col, max_row=1):
    for cell in col:
        if cell:
            if cell == 'Images URLS':
                break
            elif cell.value is not None:
                mandatory_att.append([cell.value, get_column_letter(cell.col_idx)])
r = 2
for row in ws.rows:
    # select category
    driver.find_element_by_xpath("//*[@id='choose_category']/option[contains(text(),'Car plates')]").click()
    print(mandatory_att)
    start_date_time = []
    end_date_time = []
    for att_list in mandatory_att:
        att_name = att_list[0].replace(' #', '')
        col_letter = att_list[1]
        att_value = ws[col_letter + str(r)].value
        # print(str(att_name) + ' @ ' + str(att_value))

        if att_name not in ['Car plate number', 'Start Date', 'Start Time', 'Expiration Date', 'Expiration Time',
                            'Images URLS']:
            try:
                driver.find_element_by_xpath("//form//label[contains(text(),'{}')]//following-sibling::div"
                                             "//option[contains(text(),'{}')]".format(att_name, att_value)).click()
                time.sleep(1)
            except NoSuchElementException:
                driver.find_element_by_xpath("//form//label[contains(text(),'{}')]//following-sibling::div"
                                             "/input | //form//label[contains(text(),'{}')]//following-sibling::div"
                                             "/textarea".format(att_name, att_name)).send_keys(att_value)
                time.sleep(1)
        elif att_name in ['Start Date', 'Start Time']:
            start_date_time.append(str(att_value).split(' ')[0])
            wait = WebDriverWait(driver, 100000)

            # wait.until(lambda driver: driver.current_url == 'https://mzadqatar.com/mzadqater_admin/lots/in-review-lots')
            # time.sleep(2)
            # wait.until(lambda driver: driver.current_url == 'https://mzadqatar.com/mzadqater_admin/lots/create-lot')
            x = input()
            time.sleep(2)
            print("continue" + x)

            pass
    r = r + 1
    pass
            # if len(start_date_time) == 2:
            #     st_time = datetime.strptime(start_date_time[1], "%H:%M:%S").strftime("%I:%M %p")
            #     st_date = datetime.strptime(start_date_time[0], '%Y-%m-%d').strftime('%m/%d/%Y')
            #     print([st_date, st_time])
            #     driver.find_element_by_xpath("//form//label[contains(text(),'Auction Start Time')]//following-sibling"
            #                                  "::div/input").send_keys(str('00'.join([st_date, st_time])))
            #     exit()
        # elif att_name in ['Expiration Date', 'Expiration Time']:
        #     end_date_time.append(str(att_value).split(' ')[0])
        #     if len(end_date_time) == 2:
        #         driver.find_element_by_xpath("//form//label[contains(text(),'Product Expired Time')]"
        #                                      "//following-sibling::div/input").send_keys(str('0000'.join(end_date_time)))


