from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from lxml import html
import os
import time
from datetime import datetime

executable_path = r'C:\Users\ramyg\Downloads\chromedriver_win32 (1)\chromedriver.exe'
os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
driver = webdriver.Chrome(executable_path=executable_path)
driver.get('http://shops.syaanh.com/login')

driver.find_element_by_xpath('//*[@id="email"]').send_keys('ramy.zaghloul@mzadqatar.com')
driver.find_element_by_xpath('//*[@id="password"]').send_keys('BDZuKtqUwaL2b3fd')
driver.find_element_by_xpath('//*[@id="app"]/main/div/div/div/div/div[2]/form/div[4]/div/button').click()

wb = load_workbook(filename=r'C:\Users\ramyg\Downloads\Top Meihao (1).xlsx')
ws = wb.active
i = 2
driver.switch_to.window(driver.window_handles[0])
driver.execute_script("window.open('about:blank', 'tab2');")
driver.switch_to.window("tab2")
c_skus = []


def cate(url, target_cate):
    driver.get(url)
    time.sleep(2)
    # click category tap
    driver.find_element_by_xpath('//li/a[contains(text(),"Categories")]').click()
    time.sleep(1)
    # add dropdown cate
    driver.find_element_by_xpath('//*[@id="category"]/div[1]/table/thead/tr/th[2]/div').click()
    time.sleep(1)
    # click dropdown
    driver.find_element_by_xpath('//*[@id="category"]/div/table/tbody//tr[last()]//span//button').click()
    time.sleep(1)
    # select cate
    driver.find_element_by_xpath(f'//*[@id="ui-id-1"]//li/div[contains(text(),"{target_cate}")]').click()
    time.sleep(1)
    # save
    driver.find_element_by_xpath('//nav//button[contains(text(),"Save")]').click()


for row in ws.rows:
    id = ws['a' + str(i)].value
    retail = ws['b' + str(i)].value
    cost = ws['c' + str(i)].value
    commission = ws['d' + str(i)].value
    qty = ws['e' + str(i)].value

    # driver.get('http://shops.syaanh.com/admin/mzadtopmeihao/jqadm/get/product/{}/?lang=en'.format(id))
    driver.get('http://shops.syaanh.com/admin/topmeihao/jqadm/get/product/{}/?lang=en'.format(id))
    if qty == 0:
        driver.find_element_by_xpath('//*[@id="basic"]/div[1]/div[1]/div/select/option[3]').click()
        time.sleep(1)
        driver.find_element_by_xpath('//nav//button[contains(text(),"Save")]').click()
        time.sleep(2)
    else:
        time.sleep(1)
        driver.find_element_by_xpath('//li/a[contains(text(),"Prices")]').click()
        time.sleep(1)
        driver.find_element_by_xpath('//input[@placeholder="Actual price"]').clear()
        driver.find_element_by_xpath('//input[@placeholder="Actual price"]').send_keys(str(retail))
        time.sleep(1)
        driver.find_element_by_xpath('//input[@placeholder="Cost price"]').clear()
        driver.find_element_by_xpath('//input[@placeholder="Cost price"]').send_keys(str(cost))
        time.sleep(1)
        driver.find_element_by_xpath('//input[@placeholder="Commission"]').clear()
        driver.find_element_by_xpath('//input[@placeholder="Commission"]').send_keys(str(commission))
        time.sleep(1)
        driver.find_element_by_xpath('//nav//button[contains(text(),"Save")]').click()
        time.sleep(2)
    i += 1
