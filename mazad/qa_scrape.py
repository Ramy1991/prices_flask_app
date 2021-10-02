from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from lxml import html
import os
from pywinauto.application import Application as app
from pywinauto import mouse as mouse1, keyboard
import time
from random import randrange

from datetime import datetime
import mouse

# left click


# app.menu_select('menu->Open Dashboard')
# Hubstaff.Hubstaff.menu_select('Open Dashboard')
# mouse.move(-200, 200, absolute=False, duration=0.2)
# mouse.click('left')


executable_path = r'C:\Users\ramyg\Downloads\chromedriver_win32\chromedriver.exe'
# os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
# chrome_options.add_argument('--disable-blink-features=AutomationControlled')
# chrome_options.add_argument('start-maximized')
# chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
#
# chrome_options.add_experimental_option('useAutomationExtension', False)


# chrome_options.add_argument('headless')

driver = webdriver.Chrome(options=chrome_options, executable_path=executable_path)

wb = load_workbook(filename=r"C:\Users\ramyg\Desktop\qa.xlsx")
ws = wb.active
i = 2

for row in ws.rows:
    link = ws['a' + str(i)].value
    driver.get(link)
    time.sleep(3)
    driver.find_element_by_xpath("//div[@id='page']//ul//li[8]//a[@class='glink nturl notranslate']").click()
    time.sleep(3)
    title = driver.find_element_by_xpath("//h1[@class='product_title entry-title']").text
    print(title)
    description = driver.find_element_by_xpath('//*[@id="tab-description"]').text

    ws.cell(row=i, column=2).value = ''.join(title).strip()
    ws.cell(row=i, column=3).value = ''.join(description).strip()
    wb.save(r"C:\Users\ramyg\Desktop\qa.xlsx")
    i = i + 1
