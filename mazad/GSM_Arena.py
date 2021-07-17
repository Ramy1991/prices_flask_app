import requests
from fake_useragent import UserAgent
from openpyxl import Workbook
from lxml import html
import lxml
import re
import json
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from lxml import html
import os
import time

links = [
    "https://royalford.ae/index.php?route=product/product&product_id=2075&search=RF1080WH",
    "https://royalford.ae/index.php?route=product/product&product_id=1986&search=RF1146-RT30",
    "https://royalford.ae/index.php?route=product/product&product_id=2077&search=RF1177",
    "https://royalford.ae/index.php?route=product/product&product_id=2078&search=RF1178",
    "https://royalford.ae/index.php?route=product/product&product_id=1987&search=RF1201-NSVS",
    "https://royalford.ae/index.php?route=product/product&product_id=2079&search=RF1202-NS",
    "https://royalford.ae/index.php?route=product/product&product_id=2080&search=RF1204-NSPL",
    "https://royalford.ae/index.php?route=product/product&product_id=2081&search=RF1256FP30",
    "https://royalford.ae/index.php?route=product/product&product_id=2083&search=RF1258FP18",
    "https://royalford.ae/index.php?route=product/product&product_id=1988&search=RF1259FP20",
    "https://royalford.ae/index.php?route=product/product&product_id=1989&search=RF1260FP22",
    "https://royalford.ae/index.php?route=product/product&product_id=1990&search=RF1261FP24",
    "https://royalford.ae/index.php?route=product/product&product_id=1991&search=RF1262FP26",
    "https://royalford.ae/index.php?route=product/product&product_id=1992&search=RF1263FP28",
    "https://royalford.ae/index.php?route=product/product&product_id=1993&search=RF1264FP30",
    "https://royalford.ae/index.php?route=product/product&product_id=1994&search=RF1265FP32",
    "https://royalford.ae/index.php?route=product/product&product_id=337&search=RF1270-TC",
    "https://royalford.ae/index.php?route=product/product&product_id=2084&search=RF1272-TC",
    "https://royalford.ae/index.php?route=product/product&product_id=2085&search=RF1273-TC%2FL",
    "https://royalford.ae/index.php?route=product/product&product_id=338&search=RF1274-TC",
    "https://royalford.ae/index.php?route=product/product&product_id=339&search=RF1276-TC",
    "https://royalford.ae/index.php?route=product/product&product_id=1995&search=RF1278TC",
    "https://royalford.ae/index.php?route=product/product&product_id=2086&search=RF1280-TC",
    "https://royalford.ae/index.php?route=product/product&product_id=2088&search=RF1513-IBC",
    "https://royalford.ae/index.php?route=product/product&product_id=2089&search=RF1515-IBC",
    "https://royalford.ae/index.php?route=product/product&product_id=1996&search=RF1673-S4",
    "https://royalford.ae/index.php?route=product/product&product_id=2092&search=RF1675-S6",
    "https://royalford.ae/index.php?route=product/product&product_id=2093&search=RF1676-S7",
    "https://royalford.ae/index.php?route=product/product&product_id=2094&search=RF1747-M10",
    "https://royalford.ae/index.php?route=product/product&product_id=2095&search=RF1748-M9",
    "https://royalford.ae/index.php?route=product/product&product_id=2096&search=RF1752-M10",
    "https://royalford.ae/index.php?route=product/product&product_id=2097&search=RF1796-NKT",
    "https://royalford.ae/index.php?route=product/product&product_id=1997&search=RF1916-DK",
    "https://royalford.ae/index.php?route=product/product&product_id=2099&search=RF1918-DF",
    "https://royalford.ae/index.php?route=product/product&product_id=2102&search=RF2060SP",
    "https://royalford.ae/index.php?route=product/product&product_id=1998&search=RF2064ST",
    "https://royalford.ae/index.php?route=product/product&product_id=1999&search=RF2086-CS24",
    "https://royalford.ae/index.php?route=product/product&product_id=2104&search=RF2087-CS24",
    "https://royalford.ae/index.php?route=product/product&product_id=2116&search=RF2392-TK",
    "https://royalford.ae/index.php?route=product/product&product_id=2117&search=RF2393-TBS",
    "https://royalford.ae/index.php?route=product/product&product_id=387&search=RF2394-TF",
    "https://royalford.ae/index.php?route=product/product&product_id=2002&search=RF241CBS",
    "https://royalford.ae/index.php?route=product/product&product_id=2118&search=RF2431-SP10",
    "https://royalford.ae/index.php?route=product/product&product_id=2119&search=RF2437BL",
    "https://royalford.ae/index.php?route=product/product&product_id=2003&search=RF2572",
    "https://royalford.ae/index.php?route=product/product&product_id=2121&search=RF2694-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2121&search=RF2695-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2005&search=RF2696-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2122&search=RF2700-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2006&search=RF2701-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2007&search=RF2702-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2008&search=RF2703-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2123&search=RF2704-GBD",
    "https://royalford.ae/index.php?route=product/product&product_id=2009&search=RF2763-SP",
    "https://royalford.ae/index.php?route=product/product&product_id=2126&search=RF2764-SP",
    "https://royalford.ae/index.php?route=product/product&product_id=2127&search=RF2954",
    "https://royalford.ae/index.php?route=product/product&product_id=2128&search=RF2955",
    "https://royalford.ae/index.php?route=product/product&product_id=2129&search=RF2956",
    "https://royalford.ae/index.php?route=product/product&product_id=2010&search=RF2960",
    "https://royalford.ae/index.php?route=product/product&product_id=2130&search=RF2963",
    "https://royalford.ae/index.php?route=product/product&product_id=2011&search=RF2966",
    "https://royalford.ae/index.php?route=product/product&product_id=2131&search=RF3016",
    "https://royalford.ae/index.php?route=product/product&product_id=2132&search=RF3017",
    "https://royalford.ae/index.php?route=product/product&product_id=2135&search=RF358PC7",
    "https://royalford.ae/index.php?route=product/product&product_id=2136&search=RF365IBL",
    "https://royalford.ae/index.php?route=product/product&product_id=2137&search=RF366IBM",
    "https://royalford.ae/index.php?route=product/product&product_id=2138&search=RF367IBS",
    "https://royalford.ae/index.php?route=product/product&product_id=2012&search=RF387C18",
    "https://royalford.ae/index.php?route=product/product&product_id=2013&search=RF388C20",
    "https://royalford.ae/index.php?route=product/product&product_id=2014&search=RF389C22",
    "https://royalford.ae/index.php?route=product/product&product_id=2139&search=RF390C24",
    "https://royalford.ae/index.php?route=product/product&product_id=2015&search=RF391C26",
    "https://royalford.ae/index.php?route=product/product&product_id=2140&search=RF392C28",
    "https://royalford.ae/index.php?route=product/product&product_id=2016&search=RF393C30",
    "https://royalford.ae/index.php?route=product/product&product_id=2143&search=RF398FP26",
    "https://royalford.ae/index.php?route=product/product&product_id=2018&search=RF4013",
    "https://royalford.ae/index.php?route=product/product&product_id=2146&search=RF4125PC",
    "https://royalford.ae/index.php?route=product/product&product_id=2147&search=RF4187+FK",
    "https://royalford.ae/index.php?route=product/product&product_id=2020&search=RF4188+TK",
    "https://royalford.ae/index.php?route=product/product&product_id=2148&search=RF4190+SS",
    "https://royalford.ae/index.php?route=product/product&product_id=2022&search=RF4309",
    "https://royalford.ae/index.php?route=product/product&product_id=2152&search=RF4310",
    "https://royalford.ae/index.php?route=product/product&product_id=2153&search=RF4337",
    "https://royalford.ae/index.php?route=product/product&product_id=2154&search=RF4374",
    "https://royalford.ae/index.php?route=product/product&product_id=2025&search=RF4492",
    "https://royalford.ae/index.php?route=product/product&product_id=2156&search=RF4495",
    "https://royalford.ae/index.php?route=product/product&product_id=2157&search=RF4496",
    "https://royalford.ae/index.php?route=product/product&product_id=2158&search=RF4526",
    "https://royalford.ae/index.php?route=product/product&product_id=2159&search=RF4527",
    "https://royalford.ae/index.php?route=product/product&product_id=2160&search=RF4528",
    "https://royalford.ae/index.php?route=product/product&product_id=2161&search=RF4529",
    "https://royalford.ae/index.php?route=product/product&product_id=2162&search=RF4530",
    "https://royalford.ae/index.php?route=product/product&product_id=2164&search=RF4559",
    "https://royalford.ae/index.php?route=product/product&product_id=2165&search=RF4642",
    "https://royalford.ae/index.php?route=product/product&product_id=2166&search=RF4643",
    "https://royalford.ae/index.php?route=product/product&product_id=2026&search=RF4677",
    "https://royalford.ae/index.php?route=product/product&product_id=2027&search=RF4678",
    "https://royalford.ae/index.php?route=product/product&product_id=2028&search=RF4679",
    "https://royalford.ae/index.php?route=product/product&product_id=2029&search=RF4681",
    "https://royalford.ae/index.php?route=product/product&product_id=2030&search=RF4683",
    "https://royalford.ae/index.php?route=product/product&product_id=2031&search=RF4684",
    "https://royalford.ae/index.php?route=product/product&product_id=2032&search=RF4685",
    "https://royalford.ae/index.php?route=product/product&product_id=2033&search=RF4686",
    "https://royalford.ae/index.php?route=product/product&product_id=2167&search=RF4704",
    "https://royalford.ae/index.php?route=product/product&product_id=2034&search=RF4846",
    "https://royalford.ae/index.php?route=product/product&product_id=2170&search=RF4853+SS",
    "https://royalford.ae/index.php?route=product/product&product_id=2171&search=RF4856+SP",
    "https://royalford.ae/index.php?route=product/product&product_id=2173&search=RF4860+SS",
    "https://royalford.ae/index.php?route=product/product&product_id=2174&search=RF4868+SS",
    "https://royalford.ae/index.php?route=product/product&product_id=2035&search=RF4885",
    "https://royalford.ae/index.php?route=product/product&product_id=2175&search=RF4886",
    "https://royalford.ae/index.php?route=product/product&product_id=2036&search=RF4947",
    "https://royalford.ae/index.php?route=product/product&product_id=2037&search=RF4952",
    "https://royalford.ae/index.php?route=product/product&product_id=2176&search=RF4954",
    "https://royalford.ae/index.php?route=product/product&product_id=2177&search=RF4956",
    "https://royalford.ae/index.php?route=product/product&product_id=2179&search=RF5000",
    "https://royalford.ae/index.php?route=product/product&product_id=2038&search=RF5009",
    "https://royalford.ae/index.php?route=product/product&product_id=2180&search=RF5035",
    "https://royalford.ae/index.php?route=product/product&product_id=2181&search=RF5036",
    "https://royalford.ae/index.php?route=product/product&product_id=2182&search=RF5055",
    "https://royalford.ae/index.php?route=product/product&product_id=2041&search=RF5056",
    "https://royalford.ae/index.php?route=product/product&product_id=2042&search=RF5060",
    "https://royalford.ae/index.php?route=product/product&product_id=2043&search=RF5061",
    "https://royalford.ae/index.php?route=product/product&product_id=2183&search=RF5064",
    "https://royalford.ae/index.php?route=product/product&product_id=2184&search=RF5091",
    "https://royalford.ae/index.php?route=product/product&product_id=2186&search=RF5109",
    "https://royalford.ae/index.php?route=product/product&product_id=2189&search=RF5129",
    "https://royalford.ae/index.php?route=product/product&product_id=2190&search=RF5172",
    "https://royalford.ae/index.php?route=product/product&product_id=2047&search=RF5173",
    "https://royalford.ae/index.php?route=product/product&product_id=2195&search=RF5224BL",
    "https://royalford.ae/index.php?route=product/product&product_id=2196&search=RF5224TN",
    "https://royalford.ae/index.php?route=product/product&product_id=2055&search=RF5352",
    "https://royalford.ae/index.php?route=product/product&product_id=2204&search=RF5354",
    "https://royalford.ae/index.php?route=product/product&product_id=2205&search=RF5361",
    "https://royalford.ae/index.php?route=product/product&product_id=2206&search=RF5367",
    "https://royalford.ae/index.php?route=product/product&product_id=2207&search=RF5368",
    "https://royalford.ae/index.php?route=product/product&product_id=2208&search=RF5370",
    "https://royalford.ae/index.php?route=product/product&product_id=2059&search=RF5381",
    "https://royalford.ae/index.php?route=product/product&product_id=2209&search=RF5402",
    "https://royalford.ae/index.php?route=product/product&product_id=474&search=RF5403",
    "https://royalford.ae/index.php?route=product/product&product_id=2210&search=RF5404",
    "https://royalford.ae/index.php?route=product/product&product_id=2211&search=RF5405",
    "https://royalford.ae/index.php?route=product/product&product_id=2061&search=RF5468",
    "https://royalford.ae/index.php?route=product/product&product_id=2213&search=RF5469",
    "https://royalford.ae/index.php?route=product/product&product_id=2068&search=RF5650",
    "https://royalford.ae/index.php?route=product/product&product_id=507&search=RF6304",
    "https://royalford.ae/index.php?route=product/product&product_id=552&search=RF7660",
    "https://royalford.ae/index.php?route=product/product&product_id=219&search=RF7944",
    "https://royalford.ae/index.php?route=product/product&product_id=234&search=RF8312",
    "https://royalford.ae/index.php?route=product/product&product_id=581&search=RF8325",
    "https://royalford.ae/index.php?route=product/product&product_id=582&search=RF8336",
    "https://royalford.ae/index.php?route=product/product&product_id=583&search=RF8337",
    "https://royalford.ae/index.php?route=product/product&product_id=584&search=RF8338",

]

executable_path = r"C:\Users\ramyg\Downloads\chromedriver_win32 (1)\chromedriver.exe"
os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
driver = webdriver.Chrome(executable_path=executable_path)

r = 2
wb = Workbook()
ws = wb.active

for link in links:
    driver.get(link)
    time.sleep(2)
    page = driver.page_source
    tree = html.fromstring(page)

    # item_link = tree.xpath('//*[@id="content"]/div[3]/div/div/div/div[2]/div/h4/a/@href')
    item_title = tree.xpath('//*[@id="content"]/div/div/div[2]/div/h1//text()')
    item_code = tree.xpath('//*[@id="content"]/div/div/div[2]/div/ul/li//text()')
    item_images = tree.xpath('//*[@id="content"]/div/div/div[1]/ul/li/a/img/@src')
    item_des = tree.xpath('//*[@id="content"]/div/div/div[2]/div/div[1]//text()')
    print(item_title)
    print(r)
    ws.cell(row=r, column=1).value = link
    ws.cell(row=r, column=2).value = ''.join(item_code).strip()
    ws.cell(row=r, column=3).value = ''.join(item_title).strip()
    ws.cell(row=r, column=4).value = '\n'.join(item_images).strip()
    ws.cell(row=r, column=5).value = '@'.join(item_des).strip()
    # ws.cell(row=r, column=6).value = ''.join(item_link).strip()
    r += 1
    wb.save("mi_products.xlsx")

# def user_agent():
#     return re.search(r'(.*?)\)', UserAgent().random).group(1) + "MSIE 12;)"
#
#
# header = {
#     "User-Agent": user_agent(),
#     "Accept": "*/*",
#     "Accept-Language": "*/*",
#     "Accept-Charset": "*/*",
#     "Connection": "keep-alive",
#     "Keep-Alive": "300"
# }
# r = 2
# wb = Workbook()
# ws = wb.active
# for link in links:
#     page_html = requests.get("https://royalford.ae/index.php?route=product/search&search=" + link, headers=header)
#     try:
#         tree = html.fromstring(page_html.text)
#     except lxml.etree.ParserError:
#         pass
#     des = tree.xpath('//*[@id="content"]/div[3]/div/div/div/div[2]/div/h4/a/@href')
#     # des1 =  tree.xpath('//*[@id="description"]/div/div//text()')
#     # images_links = []
#     # try:
#     #     images = re.search(r" .data..\s...thumb.(.*)", page_html.text)
#     #     json_img = json.dumps(
#     #         images.group(0).replace('],', ']').replace('"data": ', '').strip())
#     #     json_loads_images = json.loads(json.loads(json_img))
#     #
#     #     for item in json_loads_images:
#     #         images_links.append(item['full'])
#     #
#     # except AttributeError:
#     #     pass
#     # des_header = tree.xpath('//*[@id="product-details"]/div[1]/label//text() | '
#     #                         '//*[@id="product-details"]/section/dl/dt//text()')
#     # des_values = tree.xpath('//*[@id="product-details"]/div/span/a//text() | '
#     #                         '//*[@id="product-details"]/section/dl/dd//text()')
#
#     # des_dict = dict(zip(des_header, des_values))
#     # des_text = []
#     # for key in des_dict:
#     #
#     #     des_text.append(key + ": " + des_dict[key])
#     print(r)
#     print(des)
#     ws.cell(row=r, column=1).value = link
#     ws.cell(row=r, column=2).value = ''.join(des).strip()
#     r += 1
# wb.save("mi_products.xlsx")
