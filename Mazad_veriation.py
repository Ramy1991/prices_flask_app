from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from lxml import html
import os
import time
from datetime import datetime

executable_path = r'C:\Users\ramyg\Downloads\chromedriver_win32 (1)\chromedriver.exe'
os.environ['webdriver.chrome.driver'] = executable_path
chrome_options = Options()
driver = webdriver.Chrome(executable_path=executable_path)
driver.get('http://shops.syaanh.com/login')

driver.find_element_by_xpath('//*[@id="email"]').send_keys('ramy.zaghloul@mzadqatar.com')
driver.find_element_by_xpath('//*[@id="password"]').send_keys('BDZuKtqUwaL2b3fd')
driver.find_element_by_xpath('//*[@id="app"]/main/div/div/div/div/div[2]/form/div[4]/div/button').click()

wb = load_workbook(filename=r'C:\Users\ramyg\Desktop\Book2.xlsx')
ws = wb.active
i = 2
driver.switch_to.window(driver.window_handles[0])
driver.execute_script("window.open('about:blank', 'tab2');")
driver.switch_to.window("tab2")
c_skus = ["mz_63-RGS-SYQ5155", "mz_63-RGS-SYQ5154", "mz_63-RGS-SYQ5153", "mz_63-RGS-SYQ5152", "mz_63-RGS-SYQ5151", ]
for row in ws.rows:
    sku = ws['a' + str(i)].value
    p_c = ws['b' + str(i)].value
    link = ws['c' + str(i)].value
    size = ws['d' + str(i)].value

    driver.get(link)
    if p_c == 'Parent':
        driver.find_element_by_xpath('//*[@id="basic"]/div[1]/div[2]/div/select/option[5]').click()
        time.sleep(2)
        driver.find_element_by_xpath('//li/a[contains(text(),"Variants")]').click()
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="item-selection-group"]/div/div[2]/div').click()
        time.sleep(2)
        # variant
        n = 0
        for sku in c_skus:
            driver.find_element_by_xpath(
                '//*[@id="item-selection-group-data-{}"]/div[1]/div[2]/div[1]/input'.format(str(n))).send_keys(sku)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="item-selection-group"]/div/div[2]/div').click()
            n += 1
            c_skus = []
        # Texts
        element1 = driver.find_element_by_xpath('//*[@id="item-selection-group"]/div')
        driver.execute_script("return arguments[0].scrollIntoView(true);", element1)
        time.sleep(2)
        driver.find_element_by_xpath('//*[@class="col-md-3 item-navbar"]//li/a[contains(text(),"Texts")]').click()
        time.sleep(2)
        page = driver.page_source
        tree = html.fromstring(page)
        items = tree.xpath('//*[@id="item-text-group"]/div[1]/div[1]/div/div/span//text()')
        print(items)
        for item in range(len(items) - 2):
            driver.find_element_by_xpath('//*[@id="item-text-group-item-2"]/div[2]/div[3]').click()
            time.sleep(2)
        # Categories
        driver.find_element_by_xpath('//li/a[contains(text(),"Characteristics")]').click()
        time.sleep(2)
        page = driver.page_source
        tree = html.fromstring(page)
        items2 = tree.xpath('//*[@id="characteristic"]/div[3]/table/tbody//tr/td[1]//select/option/text()')
        for item2 in items2:
            if item2.strip() != 'Sold By':
                driver.find_element_by_xpath(
                    '//div[3]//tbody//option[contains(text(),"{}")]/ancestor::tr/td[@class="actions"]/div[2]'.format(
                        item2.strip())).click()
                time.sleep(2)
        # Properties

        el2 = driver.find_element_by_xpath(
            '//tr[1]//select/option[contains(text(),"SKU")]/ancestor::tbody/tr[2]//td[@class="actions"]/div')
        driver.execute_script("arguments[0].click();", el2)
        time.sleep(1)

        el1 = driver.find_element_by_xpath(
            '//tr[1]//select/option[contains(text(),"SKU")]/ancestor::tr//td[@class="actions"]/div')
        driver.execute_script("arguments[0].click();", el1)
        time.sleep(1)

        # driver.find_element_by_xpath('//nav//button[contains(text(),"Save")]').click()
    elif p_c == 'Kids':
        c_skus.append(sku)
        time.sleep(3)
        driver.find_element_by_xpath('//li/a[contains(text(),"Categories")]').click()
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="category"]/div[1]/table/tbody/tr/td[2]/div').click()
        time.sleep(2)
        driver.find_element_by_xpath('//li/a[contains(text(),"Characteristics")]').click()
        time.sleep(2)
        driver.find_element_by_xpath(
            '//*[@id="characteristic"]/div[3]//option[contains(text(),"Sold By")]/ancestor::td/following-sibling::td[2]/div[2]').click()
        time.sleep(2)
        driver.find_element_by_xpath(
            ' //*[@id="characteristic"]/div[3]//option[contains(text(),"Clothing Size")]/ancestor::td/following-sibling::td[2]/div[2]').click()
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="characteristic"]/div[2]/table/thead/tr/th[3]/div').click()
        time.sleep(2)
        driver.find_element_by_xpath(
            '//*[@id="characteristic"]/div[2]//select//option[contains(text(),"Clothing Size")]').click()
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="characteristic"]/div[2]/table/tbody/tr/td[2]/span/input').send_keys(size)
        time.sleep(2)
        driver.find_element_by_xpath('//nav//button[contains(text(),"Save")]').click()
        i += 1
